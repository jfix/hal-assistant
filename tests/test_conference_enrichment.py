from pathlib import Path

from openpyxl import load_workbook

from hal_assistant.conference_enrichment import (
    conference_enrichment_queue,
    export_conference_queue,
    import_conference_reviews,
)
from hal_assistant.models import Publication, PublicationType


def make_comm(**overrides: object) -> Publication:
    values: dict[str, object] = {
        "publication_type": PublicationType.CONFERENCE_PAPER,
        "section": "Communication dans un congrès",
        "raw_citation": "« Une communication », [colloque à Rouen, 2022].",
        "title": "Une communication",
        "year": 2024,
        "source_paragraph": 10,
        "conference_title": "Un colloque",
        "conference_city": "Rouen",
        "conference_country": "France",
    }
    values.update(overrides)
    return Publication(**values)


def test_publication_id_is_stable_across_paragraph_changes() -> None:
    first = make_comm(source_paragraph=10)
    second = make_comm(source_paragraph=99)

    assert first.publication_id == second.publication_id
    assert first.publication_id.startswith("pub-")


def test_queue_keeps_exact_date_missing_and_builds_source_queries() -> None:
    publication = make_comm()

    queue = conference_enrichment_queue([publication])

    assert len(queue) == 1
    assert queue[0]["publication_id"] == publication.publication_id
    assert queue[0]["missing_fields"] == ["conference_start_date"]
    assert any("Fabula" in query for query in queue[0]["search_queries"])


def test_ready_comm_is_not_queued() -> None:
    publication = make_comm(conference_start_date="2022-05-12")

    assert conference_enrichment_queue([publication]) == []


def test_export_creates_review_workbook(tmp_path: Path) -> None:
    json_path, workbook_path = export_conference_queue([make_comm()], tmp_path)

    assert json_path.exists()
    assert workbook_path.exists()
    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames == ["COMM review", "Instructions"]
    sheet = workbook["COMM review"]
    assert sheet["A2"].value.startswith("pub-")
    assert sheet["M2"].value == "pending"


def test_import_applies_only_accepted_sourced_metadata(tmp_path: Path) -> None:
    publication = make_comm()
    _, workbook_path = export_conference_queue([publication], tmp_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["COMM review"]
    sheet["E2"] = "2022-05-12"
    sheet["F2"] = "2022-05-13"
    sheet["J2"] = "https://www.fabula.org/actualites/example"
    sheet["K2"] = "Fabula.org"
    sheet["L2"] = "high"
    sheet["M2"] = "accepted"
    sheet["N2"] = "Programme gives both dates."
    workbook.save(workbook_path)

    publications, errors = import_conference_reviews([publication], workbook_path)

    assert errors == []
    assert publications[0].conference_start_date == "2022-05-12"
    assert publications[0].conference_end_date == "2022-05-13"
    assert {evidence.field for evidence in publications[0].metadata_evidence} == {
        "conference_start_date",
        "conference_end_date",
    }
    assert all(
        evidence.source_name == "Fabula.org"
        for evidence in publications[0].metadata_evidence
    )


def test_import_rejects_bare_year_as_conference_date(tmp_path: Path) -> None:
    publication = make_comm()
    _, workbook_path = export_conference_queue([publication], tmp_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["COMM review"]
    sheet["E2"] = "2022"
    sheet["J2"] = "https://example.org/programme"
    sheet["K2"] = "Official programme"
    sheet["L2"] = "high"
    sheet["M2"] = "accepted"
    workbook.save(workbook_path)

    publications, errors = import_conference_reviews([publication], workbook_path)

    assert publications[0].conference_start_date is None
    assert errors == ["row 2: conference_start_date must use YYYY-MM-DD"]


def test_import_ignores_nonaccepted_rows(tmp_path: Path) -> None:
    publication = make_comm()
    _, workbook_path = export_conference_queue([publication], tmp_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["COMM review"]
    sheet["E2"] = "2022-05-12"
    sheet["M2"] = "needs_review"
    workbook.save(workbook_path)

    publications, errors = import_conference_reviews([publication], workbook_path)

    assert errors == []
    assert publications[0].conference_start_date is None
