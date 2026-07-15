import json
from pathlib import Path

from openpyxl import load_workbook

from hal_assistant.exporters import export_excel, export_json
from hal_assistant.models import Enrichment, Publication, PublicationType


def sample_publication() -> Publication:
    return Publication(
        publication_type=PublicationType.JOURNAL_ARTICLE,
        section="Article dans revue",
        raw_citation="« Test », Revue, 2024, p.1-2.",
        title="Test",
        year=2024,
        pages="1-2",
        authors=["Florence Fix"],
        source_paragraph=2,
    )


def test_json_export(tmp_path: Path) -> None:
    path = export_json([sample_publication()], tmp_path / "out.json")
    data = json.loads(path.read_text(encoding="utf-8"))
    assert data[0]["title"] == "Test"


def test_excel_export(tmp_path: Path) -> None:
    path = export_excel([sample_publication()], tmp_path / "out.xlsx")
    sheet = load_workbook(path).active
    assert sheet["C2"].value == "Test"


def test_excel_export_includes_journal_authority_evidence(tmp_path: Path) -> None:
    publication = sample_publication().model_copy(
        update={
            "enrichment": Enrichment(
                source="crossref",
                score=100,
                journal="Revue validée",
                journal_id="88663",
                journal_status="VALID",
                journal_authority_score=96.5,
                issn=["1234-5678"],
                eissn=["8765-4321"],
                volume="12",
                issue="3",
                issue_title="Dossier test",
                metadata_sources=["https://example.test/authority"],
                validation_notes=["Validated against HAL journal authority 88663"],
            )
        }
    )

    path = export_excel([publication], tmp_path / "authority.xlsx")
    sheet = load_workbook(path).active
    headers = {cell.value: cell.column for cell in sheet[1]}

    assert sheet.cell(2, headers["journal authority ID"]).value == "88663"
    assert sheet.cell(2, headers["journal authority status"]).value == "VALID"
    assert sheet.cell(2, headers["journal authority score"]).value == 96.5
    assert sheet.cell(2, headers["eISSN"]).value == "8765-4321"
    assert sheet.cell(2, headers["enriched issue"]).value == "3"
    assert "authority 88663" in sheet.cell(2, headers["validation notes"]).value
